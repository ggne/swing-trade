"""
Swing Analyzer Ultimate - Multi-timeframe, Fibonacci ve Konsolidasyon Entegreli
Tam Çalışır Halde - Production Ready - Pandas 2.0+ Uyumlu
"""

import pandas as pd
import numpy as np
from tvDatafeed import TvDatafeed, Interval
import ta as ta_lib
import time, random, os, json, requests, logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from dotenv import load_dotenv
from datetime import datetime
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple
import warnings
warnings.filterwarnings('ignore')

# Thread-safe import
try:
    from advanced_utils import ErrorHandler, DataCache, ConfigValidator, SafeCalculator, ErrorSeverity
except ImportError:
    # Basit fallback implementation
    class ErrorSeverity:
        LOW = "LOW"
        MEDIUM = "MEDIUM" 
        HIGH = "HIGH"
        CRITICAL = "CRITICAL"
    
    class ErrorHandler:
        def __init__(self, max_errors=1000):
            self.error_log = []
        
        def log_error(self, message: str, severity: ErrorSeverity, symbol: str = "", function: str = ""):
            logging.error(f"{symbol} - {function}: {message}")

logger = logging.getLogger(__name__)

def setup_logging(log_file='swing_hunter_ultimate.log'):
    """Gelişmiş loglama"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, mode='w', encoding='utf-8'),
            logging.StreamHandler()
        ]
    )

# ============================================================================
# Yardımcı Sınıflar
# ============================================================================

@dataclass
class FibonacciLevel:
    """Fibonacci seviyesi"""
    level: float  # 0.236, 0.382, vb
    price: float
    distance_pct: float  # Mevcut fiyata uzaklık %
    zone: str  # 'support', 'resistance', 'neutral'

@dataclass
class ConsolidationPattern:
    """Konsolidasyon deseni"""
    detected: bool
    duration: int  # Gün sayısı
    range_pct: float  # Range genişliği %
    breakout_type: str  # 'none', 'upward', 'downward', 'potential'
    breakout_strength: float  # 0-100 arası
    support: float
    resistance: float

@dataclass
class MultiTimeframeAnalysis:
    """Çoklu zaman dilimi analizi"""
    daily_trend: str  # 'bullish', 'bearish', 'neutral'
    weekly_trend: str
    alignment: bool  # Günlük ve haftalık uyumlu mu?
    weekly_rsi: float
    weekly_macd_positive: bool
    recommendation: str  # 'strong_buy', 'buy', 'hold', 'avoid'

@dataclass
class AdvancedSignal:
    """Gelişmiş sinyal bilgisi"""
    symbol: str
    score: float
    signal_strength: str
    fibonacci_analysis: Dict
    consolidation: ConsolidationPattern
    mtf_analysis: MultiTimeframeAnalysis
    entry_zone: Tuple[float, float]  # (min, max)
    optimal_entry: float
    stop_loss: float
    target1: float
    target2: float
    risk_reward: float


# ============================================================================
# Ana Sınıf
# ============================================================================

class SwingHunterUltimate:
    """Ultimate Swing Hunter - Tüm özellikler entegre"""
    
    def __init__(self, config_path='swing_config.json'):
        self.cfg = self.load_config(config_path)
        setup_logging(self.cfg.get("log_file", "swing_hunter_ultimate.log"))
        
        load_dotenv()
        self.telegram_token = os.getenv('TELEGRAM_BOT_TOKEN')
        self.telegram_chat_id = os.getenv('TELEGRAM_CHAT_ID')
        
        # Thread-safe TvDatafeed
        self.tv = TvDatafeed()
        self.results = {"Swing Uygun": []}
        
        # Error handling
        self.error_handler = ErrorHandler()
        
        logger.info("🚀 SwingHunterUltimate başlatıldı")
    
    def load_config(self, path):
        """Config yükleme"""
        try:
            with open(path, 'r', encoding='utf-8-sig') as f:
                config = json.load(f)
            logger.info(f"✅ Config yüklendi: {path}")
            return config
        except FileNotFoundError:
            logger.warning("Config dosyası bulunamadı, varsayılan oluşturuluyor")
            default_config = {
                "symbols": ["AKBNK", "GARAN", "THYAO", "TUPRS"],
                "exchange": "BIST",
                "lookback_bars": 250,
                "min_rsi": 30.0,
                "max_rsi": 70.0,
                "min_trend_score": 50,
                "min_relative_volume": 1.0,
                "max_daily_change_pct": 8.0,
                "create_charts": True,
                "api_delay_min": 0.1,
                "api_delay_max": 0.3,
                "max_retries": 3,
                "use_multi_timeframe": True,
                "use_fibonacci": True,
                "use_consolidation": True
            }
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(default_config, f, indent=2, ensure_ascii=False)
            return default_config
        except Exception as e:
            logger.error(f"Config yükleme hatası: {e}")
            return {}
    
    def safe_api_call(self, func, *args, **kwargs):
        """Güvenli API çağrısı - Thread-safe"""
        symbol = args[0] if args else kwargs.get('symbol', 'UNKNOWN')
        
        for attempt in range(self.cfg.get('max_retries', 3)):
            try:
                delay = random.uniform(
                    self.cfg.get('api_delay_min', 0.1), 
                    self.cfg.get('api_delay_max', 0.3)
                )
                time.sleep(delay)
                
                result = func(*args, **kwargs)
                
                if result is not None and not (isinstance(result, pd.DataFrame) and result.empty):
                    return result
                    
                logger.warning(f"API'den boş veri: {symbol}")
                
            except Exception as e:
                logger.warning(f"API hatası {symbol} (deneme {attempt + 1}): {e}")
                if attempt < self.cfg.get('max_retries', 3) - 1:
                    time.sleep(2 ** attempt)
        
        logger.error(f"API çağrısı başarısız: {symbol}")
        return None
    
    def calculate_indicators(self, df):
        """Gelişmiş indikatör hesaplama - Pandas 2.0+ Uyumlu"""
        if df is None or df.empty:
            return df
            
        try:
            df = df.copy()
            
            # Temel indikatörler
            df['EMA20'] = ta_lib.trend.ema_indicator(df['close'], window=20)
            df['EMA50'] = ta_lib.trend.ema_indicator(df['close'], window=50)
            df['EMA200'] = ta_lib.trend.ema_indicator(df['close'], window=200)
            df['RSI'] = ta_lib.momentum.rsi(df['close'], window=14)
            
            # MACD
            macd = ta_lib.trend.MACD(df['close'])
            df['MACD_Level'] = macd.macd()
            df['MACD_Signal'] = macd.macd_signal()
            df['MACD_Hist'] = macd.macd_diff()
            
            # Bollinger Bands
            bb = ta_lib.volatility.BollingerBands(df['close'], window=20)
            df['BB_Upper'] = bb.bollinger_hband()
            df['BB_Lower'] = bb.bollinger_lband()
            df['BB_Middle'] = bb.bollinger_mavg()
            df['BB_Width_Pct'] = ((df['BB_Upper'] - df['BB_Lower']) / df['BB_Middle'] * 100).fillna(0)
            
            # ATR
            df['ATR14'] = ta_lib.volatility.average_true_range(
                df['high'], df['low'], df['close'], window=14
            )
            
            # Hacim analizi
            df['Volume_10d_Avg'] = df['volume'].rolling(window=10).mean()
            df['Volume_20d_Avg'] = df['volume'].rolling(window=20).mean()
            df['Relative_Volume'] = (df['volume'] / df['Volume_20d_Avg']).fillna(1)
            
            # Değişim yüzdeleri
            df['Daily_Change_Pct'] = df['close'].pct_change(1) * 100
            df['Weekly_Change_Pct'] = df['close'].pct_change(5) * 100
            
            # ADX
            adx = ta_lib.trend.ADXIndicator(df['high'], df['low'], df['close'])
            df['ADX'] = adx.adx()
            df['DI_Plus'] = adx.adx_pos()
            df['DI_Minus'] = adx.adx_neg()
            
            # Stochastic
            stoch = ta_lib.momentum.StochasticOscillator(df['high'], df['low'], df['close'])
            df['Stoch_K'] = stoch.stoch()
            df['Stoch_D'] = stoch.stoch_signal()
            
            # NaN değerleri temizle - Pandas 2.0+ uyumlu
            df = df.ffill().bfill()
            df = df.fillna(0)
            
            return df
            
        except Exception as e:
            logger.error(f"İndikatör hesaplama hatası: {e}")
            return df
    
    # ========================================================================
    # 1. FIBONACCI RETRACEMENT ANALİZİ
    # ========================================================================
    
    def calculate_fibonacci_levels(self, df, lookback=50) -> Dict[str, FibonacciLevel]:
        """
        Fibonacci retracement seviyelerini hesapla
        """
        try:
            recent = df.tail(lookback)
            
            # Swing high ve swing low bul
            swing_high = recent['high'].max()
            swing_high_idx = recent['high'].idxmax()
            swing_low = recent['low'].min()
            swing_low_idx = recent['low'].idxmin()
            
            # Trend yönünü belirle
            if swing_high_idx > swing_low_idx:
                # Uptrend - low'dan high'a çiz
                trend = 'uptrend'
                diff = swing_high - swing_low
                base = swing_low
            else:
                # Downtrend - high'dan low'a çiz
                trend = 'downtrend'
                diff = swing_high - swing_low
                base = swing_high
            
            current_price = df['close'].iloc[-1]
            
            # Fibonacci seviyeleri
            fib_ratios = {
                0.000: 'Base',
                0.236: 'Fib 23.6%',
                0.382: 'Fib 38.2%',
                0.500: 'Fib 50%',
                0.618: 'Fib 61.8%',
                0.786: 'Fib 78.6%',
                1.000: 'Top'
            }
            
            fib_levels = {}
            
            for ratio, name in fib_ratios.items():
                if trend == 'uptrend':
                    price = base + (diff * ratio)
                else:
                    price = base - (diff * ratio)
                
                distance_pct = abs((price - current_price) / current_price * 100)
                
                # Zone belirleme
                if price > current_price * 1.02:
                    zone = 'resistance'
                elif price < current_price * 0.98:
                    zone = 'support'
                else:
                    zone = 'neutral'
                
                fib_levels[name] = FibonacciLevel(
                    level=ratio,
                    price=round(price, 2),
                    distance_pct=round(distance_pct, 2),
                    zone=zone
                )
            
            logger.debug(f"Fibonacci seviyeleri hesaplandı: {trend}, Base: {base:.2f}")
            
            return {
                'trend': trend,
                'swing_high': swing_high,
                'swing_low': swing_low,
                'current_price': current_price,
                'levels': fib_levels
            }
            
        except Exception as e:
            logger.error(f"Fibonacci hesaplama hatası: {e}")
            return {}
    
    def find_fibonacci_entry_zone(self, fib_analysis: Dict) -> Tuple[float, float, float]:
        """
        Optimal Fibonacci giriş alanını bul
        """
        try:
            if not fib_analysis or 'levels' not in fib_analysis:
                return (0, 0, 0)
            
            levels = fib_analysis['levels']
            current_price = fib_analysis['current_price']
            trend = fib_analysis['trend']
            
            if trend == 'uptrend':
                # 38.2% - 50% arası ideal giriş alanı
                fib_382 = levels.get('Fib 38.2%')
                fib_500 = levels.get('Fib 50%')
                fib_618 = levels.get('Fib 61.8%')
                
                if fib_382 and fib_500:
                    zone_min = fib_382.price
                    zone_max = fib_500.price
                    
                    # Optimal giriş: %50 seviyesi
                    optimal_entry = fib_500.price
                    
                    # Eğer fiyat zone içindeyse
                    if zone_min <= current_price <= zone_max:
                        return (current_price, zone_min, zone_max)
                    
                    return (optimal_entry, zone_min, zone_max)
            
            else:  # downtrend - kısa pozisyon için
                fib_382 = levels.get('Fib 38.2%')
                fib_500 = levels.get('Fib 50%')
                
                if fib_382 and fib_500:
                    zone_min = fib_500.price
                    zone_max = fib_382.price
                    optimal_entry = fib_500.price
                    
                    return (optimal_entry, zone_min, zone_max)
            
            return (current_price, current_price * 0.98, current_price * 1.02)
            
        except Exception as e:
            logger.error(f"Fibonacci giriş hesaplama hatası: {e}")
            return (0, 0, 0)
    
    # ========================================================================
    # 2. KONSOLİDASYON VE KIRILIM TESPİTİ
    # ========================================================================
    
    def detect_consolidation_pattern(self, df, period=20, threshold_pct=8.0) -> ConsolidationPattern:
        """
        Konsolidasyon desenini ve kırılımı tespit et
        """
        try:
            recent = df.tail(period)
            
            # Range hesapla
            high_range = recent['high'].max()
            low_range = recent['low'].min()
            mid_price = (high_range + low_range) / 2
            range_pct = (high_range - low_range) / mid_price * 100
            
            # Son fiyat ve hacim
            current = df.iloc[-1]
            prev = df.iloc[-2] if len(df) > 1 else current
            
            # Konsolidasyon tespiti
            is_consolidating = range_pct < threshold_pct
            
            # Kırılım tespiti
            breakout_type = 'none'
            breakout_strength = 0.0
            
            if is_consolidating:
                # Yukarı kırılım kontrolü
                if current['close'] > high_range:
                    breakout_type = 'upward'
                    
                    # Kırılım gücü hesapla (0-100)
                    price_strength = min(
                        ((current['close'] - high_range) / high_range * 100) * 10, 
                        40
                    )
                    volume_strength = min(
                        (current.get('Relative_Volume', 1) - 1) * 30, 
                        30
                    )
                    momentum_strength = 20 if current.get('RSI', 50) > 50 else 10
                    
                    breakout_strength = price_strength + volume_strength + momentum_strength
                
                # Aşağı kırılım kontrolü
                elif current['close'] < low_range:
                    breakout_type = 'downward'
                    breakout_strength = 30  # Kısa pozisyon için önemli
                
                # Potansiyel kırılım (henüz kırılmamış ama yakın)
                elif current['close'] > high_range * 0.98:
                    if current.get('Relative_Volume', 1) > 1.3:
                        breakout_type = 'potential_upward'
                        breakout_strength = 60
                
                elif current['close'] < low_range * 1.02:
                    if current.get('Relative_Volume', 1) > 1.3:
                        breakout_type = 'potential_downward'
                        breakout_strength = 40
            
            pattern = ConsolidationPattern(
                detected=is_consolidating,
                duration=period,
                range_pct=round(range_pct, 2),
                breakout_type=breakout_type,
                breakout_strength=round(breakout_strength, 2),
                support=round(low_range, 2),
                resistance=round(high_range, 2)
            )
            
            if is_consolidating:
                logger.info(
                    f"📦 Konsolidasyon tespit edildi: Range {range_pct:.1f}%, "
                    f"Kırılım: {breakout_type}, Güç: {breakout_strength:.0f}"
                )
            
            return pattern
            
        except Exception as e:
            logger.error(f"Konsolidasyon tespit hatası: {e}")
            return ConsolidationPattern(
                detected=False, duration=0, range_pct=0, 
                breakout_type='none', breakout_strength=0,
                support=0, resistance=0
            )
    
    # ========================================================================
    # 3. MULTI-TIMEFRAME ANALİZ
    # ========================================================================
    
    def analyze_multi_timeframe(self, symbol, exchange='BIST') -> MultiTimeframeAnalysis:
        """
        Çoklu zaman dilimi analizi (Günlük + Haftalık)
        """
        try:
            # Günlük analiz
            daily = self.safe_api_call(
                self.tv.get_hist,
                symbol=symbol,
                exchange=exchange,
                interval=Interval.in_daily,
                n_bars=100
            )
            
            if daily is None or len(daily) < 20:
                logger.warning(f"Günlük veri yetersiz: {symbol}")
                return self._empty_mtf_analysis()
            
            daily_df = self.calculate_indicators(daily)
            daily_latest = daily_df.iloc[-1]
            
            # Günlük trend
            daily_trend = self._determine_trend(daily_df, daily_latest)
            
            # Haftalık analiz
            weekly = self.safe_api_call(
                self.tv.get_hist,
                symbol=symbol,
                exchange=exchange,
                interval=Interval.in_weekly,
                n_bars=52
            )
            
            if weekly is None or len(weekly) < 10:
                logger.warning(f"Haftalık veri yetersiz: {symbol}")
                return MultiTimeframeAnalysis(
                    daily_trend=daily_trend,
                    weekly_trend='unknown',
                    alignment=False,
                    weekly_rsi=50.0,
                    weekly_macd_positive=False,
                    recommendation='neutral'
                )
            
            weekly_df = self.calculate_indicators(weekly)
            weekly_latest = weekly_df.iloc[-1]
            
            # Haftalık trend
            weekly_trend = self._determine_trend(weekly_df, weekly_latest)
            
            # Haftalık göstergeler
            weekly_rsi = weekly_latest.get('RSI', 50)
            weekly_macd_positive = (
                weekly_latest.get('MACD_Level', 0) > 
                weekly_latest.get('MACD_Signal', 0)
            )
            
            # Uyum kontrolü
            alignment = (daily_trend == weekly_trend) and (daily_trend != 'neutral')
            
            # Öneri oluştur
            recommendation = self._generate_mtf_recommendation(
                daily_trend, weekly_trend, alignment, 
                weekly_rsi, weekly_macd_positive
            )
            
            mtf = MultiTimeframeAnalysis(
                daily_trend=daily_trend,
                weekly_trend=weekly_trend,
                alignment=alignment,
                weekly_rsi=round(weekly_rsi, 1),
                weekly_macd_positive=weekly_macd_positive,
                recommendation=recommendation
            )
            
            logger.info(
                f"📊 MTF Analiz {symbol}: Günlük={daily_trend}, "
                f"Haftalık={weekly_trend}, Uyum={alignment}, Öneri={recommendation}"
            )
            
            return mtf
            
        except Exception as e:
            logger.error(f"Multi-timeframe analiz hatası {symbol}: {e}")
            return self._empty_mtf_analysis()
    
    def _determine_trend(self, df, latest) -> str:
        """Trend yönünü belirle"""
        try:
            # EMA düzeni
            ema_bullish = (
                latest['close'] > latest.get('EMA20', 0) > 
                latest.get('EMA50', 0)
            )
            
            # MACD
            macd_bullish = (
                latest.get('MACD_Level', 0) > 
                latest.get('MACD_Signal', 0)
            )
            
            # ADX ve DI
            adx = latest.get('ADX', 0)
            di_plus = latest.get('DI_Plus', 0)
            di_minus = latest.get('DI_Minus', 0)
            
            adx_bullish = (adx > 20 and di_plus > di_minus)
            
            # Toplam skor
            bullish_score = sum([ema_bullish, macd_bullish, adx_bullish])
            
            if bullish_score >= 2:
                return 'bullish'
            elif bullish_score == 0:
                return 'bearish'
            else:
                return 'neutral'
                
        except Exception as e:
            logger.error(f"Trend belirleme hatası: {e}")
            return 'neutral'
    
    def _generate_mtf_recommendation(self, daily_trend, weekly_trend, 
                                     alignment, weekly_rsi, weekly_macd) -> str:
        """MTF bazlı öneri oluştur"""
        try:
            # En güçlü: Her iki timeframe da bullish
            if daily_trend == 'bullish' and weekly_trend == 'bullish':
                if weekly_rsi < 70 and weekly_macd:
                    return 'strong_buy'
                else:
                    return 'buy'
            
            # İyi: Günlük bullish, haftalık neutral
            elif daily_trend == 'bullish' and weekly_trend == 'neutral':
                return 'buy'
            
            # Dikkatli: Günlük bullish, haftalık bearish (düzeltme riski)
            elif daily_trend == 'bullish' and weekly_trend == 'bearish':
                return 'risky_buy'
            
            # Orta: Her ikisi de neutral
            elif daily_trend == 'neutral' or weekly_trend == 'neutral':
                return 'hold'
            
            # Kaçın: Bearish sinyaller
            else:
                return 'avoid'
                
        except Exception as e:
            logger.error(f"Öneri oluşturma hatası: {e}")
            return 'hold'
    
    def _empty_mtf_analysis(self) -> MultiTimeframeAnalysis:
        """Boş MTF analizi"""
        return MultiTimeframeAnalysis(
            daily_trend='neutral',
            weekly_trend='neutral',
            alignment=False,
            weekly_rsi=50.0,
            weekly_macd_positive=False,
            recommendation='hold'
        )
    
    # ========================================================================
    # 4. GELİŞMİŞ TREND SKORU
    # ========================================================================
    
    def calculate_advanced_trend_score(self, df, latest, 
                                       fib_analysis, consolidation,
                                       mtf_analysis) -> float:
        """
        Tüm faktörleri içeren gelişmiş trend skoru
        """
        try:
            score = 0.0
            
            # 1. Temel Teknik Analiz (40 puan)
            technical_score = self._calculate_technical_score(df, latest)
            score += technical_score * 0.4
            
            # 2. Fibonacci Analiz (20 puan)
            fib_score = self._calculate_fibonacci_score(fib_analysis, latest['close'])
            score += fib_score * 0.2
            
            # 3. Konsolidasyon/Kırılım (20 puan)
            consolidation_score = self._calculate_consolidation_score(consolidation)
            score += consolidation_score * 0.2
            
            # 4. Multi-timeframe (20 puan)
            mtf_score = self._calculate_mtf_score(mtf_analysis)
            score += mtf_score * 0.2
            
            final_score = min(round(score, 1), 100.0)
            
            logger.debug(
                f"Skor Detayı - Teknik: {technical_score:.0f}, "
                f"Fib: {fib_score:.0f}, Kons: {consolidation_score:.0f}, "
                f"MTF: {mtf_score:.0f}, Toplam: {final_score:.0f}"
            )
            
            return final_score
            
        except Exception as e:
            logger.error(f"Gelişmiş trend skoru hatası: {e}")
            return 50.0
    
    def _calculate_technical_score(self, df, latest) -> float:
        """Temel teknik skor (0-100)"""
        score = 0.0
        
        # EMA düzeni (25 puan)
        if latest['close'] > latest.get('EMA20', 0) > latest.get('EMA50', 0):
            score += 25
        elif latest['close'] > latest.get('EMA20', 0):
            score += 15
        
        # RSI (20 puan)
        rsi = latest.get('RSI', 50)
        if 40 <= rsi <= 60:
            score += 20
        elif 35 <= rsi <= 65:
            score += 15
        elif 30 <= rsi <= 70:
            score += 10
        
        # MACD (20 puan)
        if latest.get('MACD_Level', 0) > latest.get('MACD_Signal', 0):
            score += 15
            if len(df) > 1 and df['MACD_Hist'].iloc[-1] > df['MACD_Hist'].iloc[-2]:
                score += 5
        
        # Hacim (20 puan)
        rel_vol = latest.get('Relative_Volume', 1)
        if rel_vol >= 1.5:
            score += 20
        elif rel_vol >= 1.2:
            score += 15
        elif rel_vol >= 1.0:
            score += 10
        
        # ADX (15 puan)
        adx = latest.get('ADX', 0)
        if adx > 25:
            score += 15
        elif adx > 20:
            score += 10
        
        return min(score, 100)
    
    def _calculate_fibonacci_score(self, fib_analysis, current_price) -> float:
        """Fibonacci pozisyon skoru (0-100)"""
        if not fib_analysis or 'levels' not in fib_analysis:
            return 50  # Nötr
        
        score = 0.0
        levels = fib_analysis['levels']
        trend = fib_analysis.get('trend', 'neutral')
        
        # İdeal giriş zonunda mı?
        fib_382 = levels.get('Fib 38.2%')
        fib_500 = levels.get('Fib 50%')
        fib_618 = levels.get('Fib 61.8%')
        
        if trend == 'uptrend':
            # 38.2% - 50% arası mükemmel
            if fib_382 and fib_500:
                if fib_382.price <= current_price <= fib_500.price:
                    score = 100
                elif fib_500 and fib_618:
                    if fib_500.price <= current_price <= fib_618.price:
                        score = 80
                    else:
                        score = 60
            
            # 0% - 23.6% arası erken (riski)
            fib_base = levels.get('Base')
            fib_236 = levels.get('Fib 23.6%')
            if fib_base and fib_236:
                if fib_base.price <= current_price <= fib_236.price:
                    score = 40
            
            # 78.6% - 100% arası geç
            fib_786 = levels.get('Fib 78.6%')
            fib_100 = levels.get('Top')
            if fib_786 and fib_100:
                if fib_786.price <= current_price <= fib_100.price:
                    score = 30
        
        return score
    
    def _calculate_consolidation_score(self, consolidation: ConsolidationPattern) -> float:
        """Konsolidasyon/Kırılım skoru (0-100)"""
        if not consolidation.detected:
            return 50  # Konsolidasyon yok, nötr
        
        score = 0.0
        
        # Kırılım tipi
        if consolidation.breakout_type == 'upward':
            score = 100  # En iyi - kırılım gerçekleşti
        elif consolidation.breakout_type == 'potential_upward':
            score = 80  # İyi - kırılım yakın
        elif consolidation.breakout_type == 'none':
            score = 60  # Orta - henüz kırılım yok ama konsolidasyon var
        elif consolidation.breakout_type == 'downward':
            score = 20  # Kötü - aşağı kırılım
        else:
            score = 50
        
        # Kırılım gücünü ekle
        score = (score + consolidation.breakout_strength) / 2
        
        return min(score, 100)
    
    def _calculate_mtf_score(self, mtf_analysis: MultiTimeframeAnalysis) -> float:
        """Multi-timeframe skoru (0-100)"""
        score = 0.0
        
        # Öneri bazlı
        recommendation_scores = {
            'strong_buy': 100,
            'buy': 80,
            'risky_buy': 60,
            'hold': 50,
            'avoid': 20
        }
        
        score = recommendation_scores.get(mtf_analysis.recommendation, 50)
        
        # Uyum bonusu
        if mtf_analysis.alignment:
            score = min(score + 10, 100)
        
        return score
    
    # ========================================================================
    # 5. ANA İŞLEMLER
    # ========================================================================
    
    def process_symbol_advanced(self, symbol) -> Optional[AdvancedSignal]:
        """
        Gelişmiş sembol analizi - Tüm özellikler entegre
        """
        try:
            logger.info(f"🔍 {symbol} analiz ediliyor...")
            
            # 1. Günlük veri çek
            daily = self.safe_api_call(
                self.tv.get_hist,
                symbol=symbol,
                exchange=self.cfg.get('exchange', 'BIST'),
                interval=Interval.in_daily,
                n_bars=self.cfg.get('lookback_bars', 250)
            )
            
            if daily is None or len(daily) < 50:
                logger.warning(f"⚠️ {symbol}: Yetersiz veri")
                return None
            
            # 2. İndikatörleri hesapla
            df = self.calculate_indicators(daily)
            if df is None or df.empty:
                return None
            
            latest = df.iloc[-1]
            
            # 3. Temel filtreler
            if not self._basic_filters(latest):
                return None
            
            # 4. Fibonacci analizi
            fib_analysis = {}
            if self.cfg.get('use_fibonacci', True):
                fib_analysis = self.calculate_fibonacci_levels(df, lookback=50)
            
            # 5. Konsolidasyon tespiti
            consolidation = ConsolidationPattern(
                detected=False, duration=0, range_pct=0,
                breakout_type='none', breakout_strength=0,
                support=0, resistance=0
            )
            if self.cfg.get('use_consolidation', True):
                consolidation = self.detect_consolidation_pattern(df, period=20)
            
            # 6. Multi-timeframe analizi
            mtf_analysis = self._empty_mtf_analysis()
            if self.cfg.get('use_multi_timeframe', True):
                mtf_analysis = self.analyze_multi_timeframe(
                    symbol, 
                    self.cfg.get('exchange', 'BIST')
                )
            
            # 7. Gelişmiş trend skoru
            advanced_score = self.calculate_advanced_trend_score(
                df, latest, fib_analysis, consolidation, mtf_analysis
            )
            
            # 8. Skor filtreleme
            min_score = self.cfg.get('min_trend_score', 50)
            if advanced_score < min_score:
                logger.info(f"❌ {symbol}: Düşük skor ({advanced_score:.0f}/{min_score})")
                return None
            
            # 9. MTF öneri kontrolü
            if mtf_analysis.recommendation in ['avoid']:
                logger.info(f"❌ {symbol}: MTF öneri olumsuz ({mtf_analysis.recommendation})")
                return None
            
            # 10. Giriş noktası ve risk yönetimi
            optimal_entry, zone_min, zone_max = self.find_fibonacci_entry_zone(fib_analysis)
            
            # Stop loss ve target hesaplama
            stop_loss, target1, target2 = self._calculate_stops_targets(
                df, latest, fib_analysis, consolidation
            )
            
            # Risk/Reward
            if optimal_entry > 0 and stop_loss > 0:
                risk = optimal_entry - stop_loss
                reward = target1 - optimal_entry
                risk_reward = reward / risk if risk > 0 else 0
            else:
                risk_reward = 0
            
            # Minimum R/R kontrolü
            if risk_reward < self.cfg.get('min_risk_reward_ratio', 2.0):
                logger.info(f"❌ {symbol}: Düşük R/R ({risk_reward:.1f})")
                return None
            
            # 11. Sinyal gücü belirleme
            signal_strength = self._determine_signal_strength(
                advanced_score, risk_reward, mtf_analysis, consolidation
            )
            
            # 12. Gelişmiş sinyal objesi oluştur
            signal = AdvancedSignal(
                symbol=symbol,
                score=advanced_score,
                signal_strength=signal_strength,
                fibonacci_analysis=fib_analysis,
                consolidation=consolidation,
                mtf_analysis=mtf_analysis,
                entry_zone=(zone_min, zone_max),
                optimal_entry=optimal_entry if optimal_entry > 0 else latest['close'],
                stop_loss=stop_loss,
                target1=target1,
                target2=target2,
                risk_reward=risk_reward
            )
            
            logger.info(f"✅ {symbol}: UYGUN! Skor={advanced_score:.0f}, R/R={risk_reward:.1f}")
            
            return signal
            
        except Exception as e:
            logger.error(f"❌ {symbol} işleme hatası: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _basic_filters(self, latest) -> bool:
        """Temel filtreler"""
        cfg = self.cfg
        
        # RSI
        rsi = latest.get('RSI', 50)
        if not (cfg.get('min_rsi', 30) <= rsi <= cfg.get('max_rsi', 70)):
            return False
        
        # Göreceli hacim
        rel_vol = latest.get('Relative_Volume', 0)
        if rel_vol < cfg.get('min_relative_volume', 1.0):
            return False
        
        # Günlük değişim
        daily_change = abs(latest.get('Daily_Change_Pct', 0))
        if daily_change > cfg.get('max_daily_change_pct', 8.0):
            return False
        
        return True
    
    def _calculate_stops_targets(self, df, latest, fib_analysis, consolidation):
        """Stop loss ve target hesaplama"""
        current_price = latest['close']
        atr = latest.get('ATR14', current_price * 0.01)
        
        # Stop Loss belirleme
        stop_methods = []
        
        # 1. ATR bazlı
        atr_stop = current_price - (2 * atr)
        stop_methods.append(atr_stop)
        
        # 2. Swing low
        recent_lows = df['low'].tail(20)
        swing_low = recent_lows.min()
        stop_methods.append(swing_low)
        
        # 3. Konsolidasyon desteği
        if consolidation.detected:
            stop_methods.append(consolidation.support * 0.98)
        
        # 4. Fibonacci desteği
        if fib_analysis and 'levels' in fib_analysis:
            levels = fib_analysis['levels']
            fib_618 = levels.get('Fib 61.8%')
            if fib_618 and fib_618.zone == 'support':
                stop_methods.append(fib_618.price * 0.98)
        
        # En uygun stop'u seç (en yüksek)
        stop_loss = max(stop_methods)
        
        # Stop validasyonu
        if stop_loss >= current_price * 0.92:  # %8'den fazla risk
            stop_loss = current_price * 0.92
        
        # Target hesaplama
        risk = current_price - stop_loss
        
        # Target 1: 2R
        target1 = current_price + (risk * 2)
        
        # Target 2: Fibonacci veya 3R
        target2_options = [current_price + (risk * 3)]
        
        if consolidation.detected:
            target2_options.append(consolidation.resistance * 1.02)
        
        if fib_analysis and 'levels' in fib_analysis:
            swing_high = fib_analysis.get('swing_high', 0)
            if swing_high > current_price:
                target2_options.append(swing_high * 0.98)
        
        target2 = max(target2_options)
        
        return (
            round(stop_loss, 2),
            round(target1, 2),
            round(target2, 2)
        )
    
    def _determine_signal_strength(self, score, rr_ratio, mtf_analysis, consolidation) -> str:
        """Sinyal gücünü belirle"""
        # Temel skor
        if score >= 85 and rr_ratio >= 2.5:
            strength = "🔥🔥🔥 Mükemmel"
        elif score >= 75 and rr_ratio >= 2.0:
            strength = "🔥🔥 Çok Güçlü"
        elif score >= 65 and rr_ratio >= 1.8:
            strength = "🔥 Güçlü"
        elif score >= 55:
            strength = "⚡ Orta"
        else:
            strength = "⚠️ Zayıf"
        
        # MTF bonusu
        if mtf_analysis.recommendation == 'strong_buy':
            strength = "🔥🔥🔥 " + strength.split(' ', 1)[-1]
        
        # Kırılım bonusu
        if consolidation.breakout_type == 'upward':
            strength += " + Kırılım"
        
        return strength
    
    def run_advanced_scan(self, symbols, progress_callback=None):
        """Gelişmiş tarama"""
        self.results = {"Swing Uygun": []}
        total = len(symbols)
        
        logger.info(f"🚀 Gelişmiş tarama başlıyor: {total} sembol")
        start_time = time.time()
        
        for idx, symbol in enumerate(symbols):
            if progress_callback:
                progress = int(((idx + 1) / total) * 100)
                progress_callback(
                    progress, 
                    f"{idx+1}/{total} - {symbol} analiz ediliyor..."
                )
            
            signal = self.process_symbol_advanced(symbol)
            
            if signal:
                # Excel için format
                result = self._format_result(signal)
                self.results["Swing Uygun"].append(result)
        
        # Sonuçları skor ve MTF'ye göre sırala
        if self.results["Swing Uygun"]:
            self.results["Swing Uygun"].sort(
                key=lambda x: (
                    1 if x.get('MTF Öneri') == 'strong_buy' else 0,
                    float(x.get('Skor', 0))
                ),
                reverse=True
            )
        
        elapsed = time.time() - start_time
        logger.info(
            f"✅ Tarama tamamlandı: {len(self.results['Swing Uygun'])} sonuç, "
            f"{elapsed:.1f} saniye"
        )
        
        return self.results
    
    def _format_result(self, signal: AdvancedSignal) -> Dict:
        """Sinyal objesini Excel formatına çevir"""
        latest_price = signal.optimal_entry
        
        # Fibonacci bilgisi
        fib_info = "Yok"
        if signal.fibonacci_analysis and 'levels' in signal.fibonacci_analysis:
            fib_trend = signal.fibonacci_analysis.get('trend', 'unknown')
            fib_info = f"{fib_trend.title()}"
        
        # Konsolidasyon bilgisi
        cons_info = "Yok"
        if signal.consolidation.detected:
            cons_info = f"{signal.consolidation.breakout_type.title()} ({signal.consolidation.range_pct:.1f}%)"
        
        return {
            'Hisse': signal.symbol,
            'Fiyat': f"{latest_price:.2f}",
            'Sinyal': signal.signal_strength,
            'Skor': f"{signal.score:.0f}/100",
            'Giriş (Min-Max)': f"{signal.entry_zone[0]:.2f} - {signal.entry_zone[1]:.2f}",
            'Optimal Giriş': f"{signal.optimal_entry:.2f}",
            'Stop Loss': f"{signal.stop_loss:.2f}",
            'Hedef 1': f"{signal.target1:.2f}",
            'Hedef 2': f"{signal.target2:.2f}",
            'R/R': f"1:{signal.risk_reward:.1f}",
            'Risk %': f"{((signal.optimal_entry - signal.stop_loss) / signal.optimal_entry * 100):.1f}",
            'Günlük Trend': signal.mtf_analysis.daily_trend.title(),
            'Haftalık Trend': signal.mtf_analysis.weekly_trend.title(),
            'MTF Uyum': '✅' if signal.mtf_analysis.alignment else '❌',
            'MTF Öneri': signal.mtf_analysis.recommendation.replace('_', ' ').title(),
            'Fibonacci': fib_info,
            'Konsolidasyon': cons_info,
        }
    
    def save_to_excel(self, results):
        """Gelişmiş Excel raporu"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Swing_Ultimate_Raporu_{timestamp}.xlsx"
        
        try:
            df = pd.DataFrame(results.get("Swing Uygun", []))
            
            if df.empty:
                logger.warning("Rapor oluşturulacak uygun hisse bulunamadı.")
                return None
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name="Swing Signals", index=False)
                ws = writer.sheets["Swing Signals"]
                
                # Header stil
                header_fill = PatternFill("solid", fgColor="1F4E78")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Sütun genişlikleri
                column_widths = {
                    'A': 12,  # Hisse
                    'B': 10,  # Fiyat
                    'C': 25,  # Sinyal
                    'D': 12,  # Skor
                    'E': 20,  # Giriş
                    'F': 15,  # Optimal
                    'G': 12,  # Stop
                    'H': 12,  # Hedef1
                    'I': 12,  # Hedef2
                    'J': 10,  # R/R
                    'K': 10,  # Risk
                    'L': 15,  # Günlük
                    'M': 15,  # Haftalık
                    'N': 12,  # Uyum
                    'O': 20,  # Öneri
                    'P': 15,  # Fib
                    'Q': 25,  # Kons
                }
                
                for col, width in column_widths.items():
                    ws.column_dimensions[col].width = width
                
                # Skor renklendirme
                for row in range(2, len(df) + 2):
                    # Skor sütunu
                    score_cell = ws[f'D{row}']
                    try:
                        score_val = float(score_cell.value.split('/')[0])
                        
                        if score_val >= 80:
                            score_cell.fill = PatternFill("solid", fgColor="00B050")
                            score_cell.font = Font(bold=True, color="FFFFFF")
                        elif score_val >= 70:
                            score_cell.fill = PatternFill("solid", fgColor="92D050")
                        elif score_val >= 60:
                            score_cell.fill = PatternFill("solid", fgColor="FFFF00")
                    except:
                        pass
                    
                    # MTF Uyum
                    mtf_cell = ws[f'N{row}']
                    if mtf_cell.value == '✅':
                        mtf_cell.fill = PatternFill("solid", fgColor="C6EFCE")
                    
                    # MTF Öneri
                    rec_cell = ws[f'O{row}']
                    rec_value = str(rec_cell.value).lower()
                    if 'strong buy' in rec_value:
                        rec_cell.fill = PatternFill("solid", fgColor="00B050")
                        rec_cell.font = Font(bold=True, color="FFFFFF")
                    elif 'buy' in rec_value:
                        rec_cell.fill = PatternFill("solid", fgColor="92D050")
                
                # Freeze panes
                ws.freeze_panes = 'A2'
            
            logger.info(f"✅ Rapor başarıyla oluşturuldu: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"Excel kaydetme hatası: {e}")
            return None

# ============================================================================
# TEST FONKSİYONU
# ============================================================================

def test_ultimate_scanner():
    """Ultimate scanner testi"""
    print("\n🚀 SwingHunterUltimate Test Başlıyor...")
    print("="*70)
    
    try:
        # Hunter oluştur
        hunter = SwingHunterUltimate('swing_config.json')
        
        # Test sembolleri
        test_symbols = ['GARAN', 'AKBNK', 'THYAO', 'TUPRS']
        
        print(f"\n📊 Test sembolleri: {test_symbols}")
        print("\n🔍 Tarama başlıyor...")
        
        # Tarama yap
        results = hunter.run_advanced_scan(test_symbols)
        
        print(f"\n✅ Tarama tamamlandı!")
        print(f"📈 Bulunan hisse: {len(results['Swing Uygun'])}")
        
        # Sonuçları göster
        for idx, stock in enumerate(results['Swing Uygun'], 1):
            print(f"\n{idx}. 🎯 {stock['Hisse']}")
            print(f"   Sinyal: {stock['Sinyal']}")
            print(f"   Skor: {stock['Skor']}")
            print(f"   R/R: {stock['R/R']}")
            print(f"   MTF: {stock['Günlük Trend']} / {stock['Haftalık Trend']}")
            print(f"   Öneri: {stock['MTF Öneri']}")
        
        # Excel rapor
        excel_file = hunter.save_to_excel(results)
        if excel_file:
            print(f"\n📊 Excel raporu: {excel_file}")
        
        print("\n✅ Test başarıyla tamamlandı!")
        return True
        
    except Exception as e:
        print(f"\n❌ Test hatası: {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == '__main__':
    # Test çalıştır
    success = test_ultimate_scanner()
    
    if success:
        print("\n🎉 Sistem hazır! GUI ile kullanabilirsiniz.")
        print("   python gui-ultimate-integration.py")
    else:
        print("\n⚠️ Test başarısız, lütfen hataları kontrol edin.")